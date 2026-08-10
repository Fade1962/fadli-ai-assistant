from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def create_xlsx(filename, content):
    wb = Workbook()
    ws = wb.active
    ws.title = "NARA"
    ws.append(["NARA Result"])
    ws["A1"].font = Font(bold=True)
    for line in (content or "").splitlines():
        ws.append([line])
    ws.column_dimensions["A"].width = 120
    for cell in ws["A"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(filename)
    return filename
