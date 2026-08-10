import os

import pandas as pd
from openpyxl import load_workbook


def process_spreadsheet(filename):

    try:

        extension = os.path.splitext(
            filename
        )[1].lower()


        # ==========================================
        # CSV
        # ==========================================

        if extension == ".csv":

            dataframe = pd.read_csv(
                filename
            )

            return dataframe_to_text(
                dataframe
            )


        # ==========================================
        # XLS / XLSX
        # ==========================================

        workbook = load_workbook(

            filename,

            read_only=True,

            data_only=True

        )


        output = []


        for sheet in workbook.worksheets:

            output.append(
                f"=== SHEET: {sheet.title} ==="
            )


            for row in sheet.iter_rows(
                values_only=True
            ):

                values = [

                    str(value)
                    if value is not None
                    else ""

                    for value in row

                ]


                if any(values):

                    output.append(
                        " | ".join(values)
                    )


        workbook.close()


        result = "\n".join(
            output
        ).strip()


        if not result:

            return (
                "Spreadsheet berhasil dibaca, "
                "tetapi tidak ditemukan data."
            )


        return result


    except Exception as error:

        print(
            "SPREADSHEET PROCESS ERROR:",
            repr(error)
        )

        raise



def dataframe_to_text(
    dataframe
):

    output = []


    columns = [

        str(column)

        for column in dataframe.columns

    ]


    output.append(
        " | ".join(columns)
    )


    for _, row in dataframe.iterrows():

        values = [

            str(value)
            if not pd.isna(value)
            else ""

            for value in row

        ]


        output.append(
            " | ".join(values)
        )


    return "\n".join(
        output
    ).strip()
